# 2.13.2021 - Python Youtube 2:11:55 forward
import openpyxl as xl
wb = xl.load_workbook(('transactions.xlsx'))
print(wb.sheetnames)
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)

for row in range(2, sheet.max_row + 1): # range generates 1, 2, 3 for range(1, 4), ignore the title row
    print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save('transaction2.xlsx')

# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types wil
# l automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")

# from pathlib import Path
# # Absolute path
# # C:\Program Files\Microsft
# # Relative path
#
# path = Path("ecommerce")
# print(path.exists())
#
# path = Path("ecommerce1")
# print(path.exists())
#
# path = Path("emails")
# print(path.mkdir())
# print(path.rmdir())
#
# path = Path()
# print(path.glob('*.py'))
#
# for file in path.glob('*.py'):
#     print(file)
#
# for file in path.glob('*'):
#     print(file)


# import random


# def dice():
#     dice_faces = (1, 2, 3, 4, 5, 6)
#     return random.choice(dice_faces)
#
# roll1 = dice()
# roll2 = dice()
# print(f"({roll1}, {roll2})")
# print(f"Total = {roll1 + roll2}")


# class Dice:
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         return first, second
#         # returning a tuple, don't need parenthesis
#
#     def roll8(self):
#         first = random.randint(1, 8)
#         second = random.randint(1, 8)
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())
# print(dice.roll8())

# for i in range(3):
#     print(random.random())
#     print(random.randint(10,20))

# members = ['John', 'Mary', 'Bob', 'Mosh']
# leader = random.choice(members)
#
# print(leader)

# converters.py

# import converters
#
# print(converters.kg_to_lbs(70))


# from converters import kg_to_lbs
#
# print(kg_to_lbs(100))

# numbers = [10, 3, 6, 2]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)
#
def find_max(array):
    maximum = array[0]
    for number in array:
        if number > maximum:
            maximum = number
    return maximum

# print(find_max(numbers))
#
# # import utils
# #
# # numbers = [10, 3, 6, 2]
# # print(utils.find_max(numbers))
#
# # from utils import find_max
# #
# numbers = [10, 12, 3, 6, 2]
# # print(find_max(numbers))
#
# print(max(numbers))

# IMPORTING FUNCTIONS FROM A PACKAGE/MODULE
# import ecommerce.shipping
# ecommerce.shipping.calc_shipping()
#
# from ecommerce.shipping import calc_shipping
# calc_shipping()
